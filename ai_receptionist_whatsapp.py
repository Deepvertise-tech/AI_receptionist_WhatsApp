import os, json, base64, logging, asyncio, time, struct, re, html, unicodedata
from typing import Optional, Dict, Any
from dotenv import load_dotenv
from quart import Quart, request, websocket
from twilio.twiml.voice_response import VoiceResponse
from twilio.twiml.messaging_response import MessagingResponse  # NEW: for WhatsApp replies
import azure.cognitiveservices.speech as speechsdk
import httpx

# --- EMAIL ADD (imports) ---
import smtplib
from email.message import EmailMessage
# ---------------------------

# ---- NEW: Excel (OpenPyXL) support ----
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# ---------- Setup ----------
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("receptionist")

app = Quart(__name__)

# ---------- μ-law helpers ----------
MU_LAW_BIAS = 0x84
EXP_LUT = [0,132,396,924,1980,4092,8316,16764]
def mulaw_byte_to_pcm16(mu):
    mu = ~mu & 0xFF
    sign = (mu & 0x80)
    exponent = (mu >> 4) & 0x07
    mantissa = mu & 0x0F
    magnitude = EXP_LUT[exponent] + (mantissa << (exponent + 3))
    sample = (magnitude - MU_LAW_BIAS)
    if sign:
        sample = -sample
    return max(-32768, min(32767, sample))

def mulaw_b64_to_pcm16_bytes(b64):
    raw = base64.b64decode(b64)
    out = bytearray()
    for b in raw:
        s = mulaw_byte_to_pcm16(b)
        out += int(s).to_bytes(2, "little", signed=True)
    return bytes(out)

# ---------- Simple VAD (adaptive) ----------
_noise_floor = {"ema": 300.0}
def is_speech(pcm16: bytes, margin=2.6):
    if not pcm16:
        return False
    samples = struct.unpack("<" + "h"*(len(pcm16)//2), pcm16)
    avg = sum(abs(s) for s in samples) / max(1, len(samples))
    ema = _noise_floor["ema"] = 0.95 * _noise_floor["ema"] + 0.05 * avg
    thresh = ema * margin + 180
    return avg > thresh

# ---------- LLM client (HTTP/2) ----------
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
httpx_timeout = httpx.Timeout(connect=4.0, read=20.0, write=10.0, pool=4.0)
httpx_client = httpx.AsyncClient(http2=True, timeout=httpx_timeout, headers={
    "Authorization": f"Bearer {OPENAI_API_KEY}",
})

# ---------- Env helpers (UTF-8 safe) ----------
def _env_text(name: str, default: Optional[str] = None) -> Optional[str]:
    """
    Read a UTF-8 env var safely (keeps umlauts etc.). Only unescape \n and \t.
    """
    val = os.getenv(name)
    if val is None or val == "":
        return default
    return val.replace("\\n", "\n").replace("\\t", "\t")

def _parse_env_list(var_name: str):
    """
    Accept JSON array or comma-separated list from .env.
    Returns list[str]. Missing or empty -> [].
    """
    val = os.getenv(var_name, "").strip()
    if not val:
        return []
    if val.startswith("["):
        try:
            arr = json.loads(val)
            return [str(x).strip() for x in arr if str(x).strip()]
        except Exception:
            logger.warning(f"[ENV] {var_name} JSON parse failed; falling back to CSV split.")
    return [x.strip() for x in val.split(",") if x.strip()]

# ---------- Core ----------
class LowLatencyReceptionist:
    def __init__(self):
        self.business_info = self._load_business_info()

        # Optional clamp to avoid very large prompts hurting latency
        MAX_BIZ = int(os.getenv("BUSINESS_INFO_MAX_CHARS", "12000"))
        if len(self.business_info) > MAX_BIZ:
            logger.warning(f"[PROMPT] BUSINESS_INFO truncated from {len(self.business_info)} to {MAX_BIZ} chars")
            self.business_info = self.business_info[:MAX_BIZ]

        # BASE_SYSTEM_PROMPT comes ONLY from env; expand placeholders
        raw = _env_text("BASE_SYSTEM_PROMPT", "") or ""
        raw = raw.replace("{BUSINESS_INFO}", self.business_info)
        raw = raw.replace("{COMPANY_NAME}", self._company_name())

        # ---- MINIMAL ADD: ensure explicit recall permission/instruction ----
        raw += (
            "\n\n### Data Recall & Recap Rules\n"
            "- If the user asks to **repeat/recap/show** the stored details, you MUST provide a **complete summary** of all known caller and booking info.\n"
            "- It is permitted to repeat **the caller’s own details** (name, phone, email, address, and appointment specifics) when they ask. Do not refuse.\n"
            "- Format the recap as a short bullet list.\n"
        )
        # ---------------------------------------------------------------

        self.base_system_prompt = raw
        logger.info(f"[PROMPT] base_system_prompt chars={len(self.base_system_prompt)}")

    def _load_business_info(self):
        fn = os.getenv("BUSINESS_INFO_FILE", "business_info.txt")
        try:
            with open(fn, "r", encoding="utf-8") as f:
                return f.read()
        except:
            return "Business information not available."

    def _company_name(self):
        for line in self.business_info.splitlines():
            if "Company Name:" in line:
                return line.split("Company Name:")[1].strip()
        return "our company"

    def make_asr(self):
        key, region = os.getenv("AZURE_SPEECH_KEY"), os.getenv("AZURE_SPEECH_REGION")
        if not key or not region:
            raise RuntimeError("Missing AZURE_SPEECH_KEY / AZURE_SPEECH_REGION")
        speech_config = speechsdk.SpeechConfig(subscription=key, region=region)
        speech_config.speech_recognition_language = "de-DE"
        speech_config.set_property(speechsdk.PropertyId.SpeechServiceConnection_EndSilenceTimeoutMs, "200")
        speech_config.set_property(speechsdk.PropertyId.SpeechServiceConnection_InitialSilenceTimeoutMs, "5000")
        fmt = speechsdk.audio.AudioStreamFormat(samples_per_second=8000, bits_per_sample=16, channels=1)
        push_stream = speechsdk.audio.PushAudioInputStream(fmt)
        audio_in = speechsdk.audio.AudioConfig(stream=push_stream)
        recognizer = speechsdk.SpeechRecognizer(speech_config=speech_config, audio_config=audio_in)

        # CUSTOM_PHRASES ONLY from env
        try:
            phrases = _parse_env_list("CUSTOM_PHRASES")
            logger.info(f"[ASR] Loaded {len(phrases)} custom phrases: {phrases[:10]}")
            if phrases:
                phrase_list = speechsdk.PhraseListGrammar.from_recognizer(recognizer)
                for p in phrases:
                    phrase_list.addPhrase(p)
        except Exception as e:
            logger.warning(f"Could not add custom phrases to ASR: {e}")
        return recognizer, push_stream

    def make_tts(self):
        key, region = os.getenv("AZURE_SPEECH_KEY"), os.getenv("AZURE_SPEECH_REGION")
        if not key or not region:
            raise RuntimeError("Missing AZURE_SPEECH_KEY / AZURE_SPEECH_REGION")
        speech_config = speechsdk.SpeechConfig(subscription=key, region=region)
        speech_config.set_speech_synthesis_output_format(
            speechsdk.SpeechSynthesisOutputFormat.Raw8Khz8BitMonoMULaw
        )
        speech_config.speech_synthesis_voice_name = os.getenv("AZURE_TTS_VOICE", "de-DE-AmalaNeural")
        synthesizer = speechsdk.SpeechSynthesizer(speech_config=speech_config, audio_config=None)
        return synthesizer

state = LowLatencyReceptionist()

# ---------- Contact & address extraction ----------
INTL_PHONE_CANDIDATE = re.compile(r'(\+?\d[\d\s().-]{2,}\d)')
_NUM_WORD = {"zero":"0","oh":"0","o":"0","one":"1","two":"2","three":"3","four":"4","five":"5","six":"6","seven":"7","eight":"8","nine":"9"}
_SEP_WORDS = {"dash","hyphen","space","dot","point"}
_PLUS_WORDS = {"plus"}
_REPEAT_WORDS = {"double": 2, "triple": 3}

# Support EN + DE name statements
NAME_PATTERNS = [
    re.compile(r"\bmy name is\s+([A-Za-zÄÖÜäöüß][A-Za-zÄÖÜäöüß'\-]+(?:\s+[A-Za-zÄÖÜäöüß][A-Za-zÄÖÜäöüß'\-]+){0,2})\b(?!\s*(?:and|my|number|is|phone|contact|,|\.))", re.I),
    re.compile(r"\bthis is\s+([A-Za-zÄÖÜäöüß][A-Za-zÄÖÜäöüß'\-]+(?:\s+[A-Za-zÄÖÜäöüß][A-Za-zÄÖÜäöüß'\-]+){0,2})\b(?!\s*(?:and|my|number|is|phone|contact|,|\.))", re.I),
    re.compile(r"\bI am\s+([A-Za-zÄÖÜäöüß][A-Za-zÄÖÜäöüß'\-]+(?:\s+[A-Za-zÄÖÜäöüß][A-Za-zÄÖÜäöüß'\-]+){0,2})\b(?!\s*(?:and|my|number|is|phone|contact|,|\.))", re.I),
    re.compile(r"\b(?:mein name ist|ich hei(?:s|ß)e)\s+([A-Za-zÄÖÜäöüß][A-Za-zÄÖÜäöüß'\-]+(?:\s+[A-Za-zÄÖÜäöüß][A-Za-zÄÖÜäöüß'\-]+){0,2})\b", re.I),
]

PLZ_RE = re.compile(r'\b(?:D-)?(?P<plz>\d{5})\b')
_STREET_SUFFIX_ALT = r'(straße|strasse|str\.|weg|allee|platz|ring|gasse|ufer|damm|kai|markt|stieg|steig|pfad|chaussee)'
STREET_EMBED_RE = re.compile(rf'\b(?P<street>[A-ZÄÖÜ][A-Za-zÄÖÜäöüß\-.]*?(?:{_STREET_SUFFIX_ALT}))\b', re.IGNORECASE)
STREET_SPACED_RE = re.compile(rf'\b(?P<street>(?:[A-ZÄÖÜ][A-Za-zÄÖÜäöüß\-.]+(?:\s+[A-ZÄÖÜ][A-Za-zÄÖÜäöüß\-.]+)*)\s+{_STREET_SUFFIX_ALT})\b', re.IGNORECASE)
HOUSE_AFTER_RE = re.compile(r'\s+(?P<house>\d+[A-Za-z]?(?:-\d+[A-Za-z]?)?)(?!\d)')
EMAIL_TEXT_RE = re.compile(r'\b([A-Za-z0-9._%+\-]+)@([A-Za-z0-9.\-]+\.[A-Za-z]{2,})\b')

# Include German phone cues
_PHONE_CUE = re.compile(r"(my|the|a|meine|mein)?\s*(phone|number|mobile|cell|contact|reach me|call me|callback|call\-back|telefonnummer|rufnummer|handy|nummer)\s*(is|ist|:)?", re.I)

_STREET_SUFFIX = r'(straße|strasse|str\.|weg|allee|platz|ring|gasse|ufer|damm|kai|markt|stieg|steig|pfad|chaussee)'
STREET_RE = re.compile(rf'\b(?P<street>[A-ZÄÖÜ][A-Za-zÄÖÜäöüß\-.]+(?:\s+[A-ZÄÖÜ][A-Za-zÄÖÜäöüß\-.]+)*\s+{_STREET_SUFFIX})\s+(?P<house>\d+[A-Za-z]?(\-\d+[A-Za-z]?)?)\b')
PLZ_CITY_RE = re.compile(r'\b(?:D-)?(?P<plz>\d{5})\s*[, ]\s*(?P<city>[A-Za-zÄÖÜäöüß\-.]+(?:\s+[A-Za-zÄÖÜäöüß\-.]+)*)\b', re.IGNORECASE)

# --- NEW: lightweight update intent detection (EN + DE) ---
_UPDATE_CUE = re.compile(
    r"\b(update|change|correct|wrong|actually|new|neu|ändern|geändert|korrigier|korrektur|richtig|falsch|statt|jetzt|neue[rn]?|neues)\b",
    re.IGNORECASE
)
def _wants_update(text: str) -> bool:
    t = (text or "")
    if _UPDATE_CUE.search(t):
        return True
    # direct declaratives often imply setting/overwriting
    return bool(re.search(r"(my name is|mein name ist|ich hei(?:s|ß)e|my number is|meine (?:telefon)?nummer ist|my email is|meine e[\-\s]?mail ist|my address is|meine adresse ist)", t, re.IGNORECASE))

def _set_field(contact: dict, key: str, new_val: Optional[str], *, override: bool, label: str):
    if not new_val:
        return
    old = contact.get(key)
    if old and not override and old == new_val:
        return
    if old and not override and old != new_val:
        # If user didn't clearly signal update, keep first unless new looks clearly intended (e.g., longer)
        if len(new_val) + 1 < len(old):
            return
    contact[key] = new_val
    if old and old != new_val:
        logger.info(f"[MEM] updated {label}: {old} -> {new_val}")
    else:
        logger.info(f"[MEM] captured {label}: {new_val}")

def _norm_strasse_case(s: str) -> str:
    return re.sub(r'(?i)strasse\b', 'straße', s)

# ---- NEW: Excel append helper ----
def _append_record_excel(name: str, phone: str, path: str = "Record.xlsx"):
    if not OPENPYXL_AVAILABLE:
        logger.error("[Excel] openpyxl not available; cannot write Record.xlsx]")
        return
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            # ensure headers
            if ws.max_row == 0 or (ws.max_row == 1 and (ws.cell(row=1, column=1).value is None)):
                ws.cell(row=1, column=1, value="Name")
                ws.cell(row=1, column=2, value="Number")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.cell(row=1, column=1, value="Name")
            ws.cell(row=1, column=2, value="Number")
        ws.append([name, phone])
        wb.save(path)
        logger.info(f"[Excel] Appended to {path}: {name} | {phone}")
    except Exception as e:
        logger.error(f"[Excel] Failed to write Record.xlsx: {e}")

def extract_address(user_text: str, call_state: dict):
    if not user_text or "@" in user_text:
        return
    override = _wants_update(user_text)
    contact = call_state["contact"]
    addr = contact.setdefault("address", {"street": None, "house": None, "postal": None, "city": None})

    # If explicit update/correction – start fresh
    if override:
        addr.update({"street": None, "house": None, "postal": None, "city": None})

    txt = _norm_strasse_case(user_text.strip())

    # PLZ + optional city (any order)
    plz_match = PLZ_RE.search(txt)
    work = txt
    if plz_match:
        addr["postal"] = plz_match.group("plz")
        work = txt[:plz_match.start()] + ' ' + txt[plz_match.end():]

    plz_city = PLZ_CITY_RE.search(txt)
    if plz_city:
        addr["postal"] = plz_city.group("plz")
        addr["city"]   = plz_city.group("city").strip().title()

    # If we know postal but not city, try infer next word(s)
    if addr.get("postal") and not addr.get("city"):
        tokens = re.findall(r'[A-Za-zÄÖÜäöüß\-.]+|\d{5}', txt)
        if addr["postal"] in tokens:
            i = tokens.index(addr["postal"])
            if i + 1 < len(tokens) and tokens[i+1][0].isalpha():
                city_guess = tokens[i+1]
                j = i + 2
                while j < len(tokens) and tokens[j][0].isalpha():
                    city_guess += " " + tokens[j]; j += 1
                addr["city"] = city_guess.title()

    # Street + house
    if not addr.get("street") or override:
        m = STREET_EMBED_RE.search(work) or STREET_SPACED_RE.search(work)
        if m:
            addr["street"] = _norm_strasse_case(m.group("street")).strip()
            tail = work[m.end():]
            h = HOUSE_AFTER_RE.match(tail)
            if h:
                addr["house"] = h.group(["house"])

    # If street known but house missing, try after street
    if addr.get("street") and (not addr.get("house") or override):
        mstreet = re.search(re.escape(addr["street"]), txt, re.IGNORECASE)
        if mstreet:
            after = txt[mstreet.end():]
            mhouse = re.search(r'\b(\d+[A-Za-z]?(?:-\d+[A-Za-z]?)?)\b(?!\s*\d)', after)
            if mhouse:
                addr["house"] = mhouse.group(1)

    # Untangle house vs postal mashups
    if addr.get("house"):
        if re.fullmatch(r'\d{6,}', addr["house"]):
            if not addr.get("postal"):
                addr["postal"] = addr["house"][-5:]
            addr["house"] = addr["house"][:-5]
        m = re.fullmatch(r'(\d{1,4}[A-Za-z]?)[^\d]*?(\d{5})', addr["house"])
        if m:
            if not addr.get("postal"):
                addr["postal"] = m.group(2)
            addr["house"] = m.group(1)
    if addr.get("house") == "":
        addr["house"] = None

    logger.info(f"[MEM] address now: {addr}")

def _looks_like_phone_utterance(text: str) -> bool:
    return bool(_PHONE_CUE.search(text or ""))

def _normalize_numeric_candidate(s: str, *, allow_short: bool) -> Optional[str]:
    if not s:
        return None
    s = s.strip()
    lead_plus = s.lstrip().startswith("+")
    digits = re.sub(r"\D", "", s)
    if s.startswith("00"):
        digits = digits[2:]
        lead_plus = True
    min_len = 5 if allow_short else 7
    if not (min_len <= len(digits) <= 15):
        return None
    return ("+" if lead_plus else "") + digits

def _spoken_to_digits(text: str, *, allow_short: bool) -> Optional[str]:
    toks = re.findall(r"[a-zA-ZÄÖÜäöüß]+|\d+|\+|[\-().]", (text or "").lower())
    # include German digit words
    de_map = {
        "null":"0","eins":"1","ein":"1","zwei":"2","drei":"3","vier":"4","fünf":"5","funf":"5","sechs":"6",
        "sieben":"7","acht":"8","neun":"9","zehn":"10"
    }
    num_map = {**_NUM_WORD, **de_map}
    out, lead_plus, i = [], False, 0
    while i < len(toks):
        t = toks[i]
        if t in _PLUS_WORDS or t == "+":
            lead_plus = True
            i += 1
            continue
        if t in _SEP_WORDS or t in {"-", "(", ")", "."}:
            i += 1
            continue
        if t in _REPEAT_WORDS and (i + 1) < len(toks):
            nxt = toks[i+1]
            d = num_map.get(nxt)
            if d and len(d) == 1:
                out.extend(d * _REPEAT_WORDS[t])
                i += 2
                continue
        d = num_map.get(t)
        if d:
            out.extend(list(d))
            i += 1
            continue
        if t.isdigit():
            out.extend(list(t))
            i += 1
            continue
        i += 1
    digits = "".join(out)
    min_len = 5 if allow_short else 7
    if min_len <= len(digits) <= 15:
        return ("+" if lead_plus else "") + digits
    return None

def _clean_name(name: str) -> str:
    name = re.split(r"\b(?:and|my|number|is|phone|contact|mein|meine|nummer|telefonnummer)\b|[,\.]", name, 1, flags=re.I)[0].strip()
    parts = [p for p in name.split() if p][:3]
    return " ".join(w[0:1].upper() + w[1:] for w in parts)

def extract_contact(user_text: str, call_state: dict):
    if not user_text:
        return
    contact = call_state["contact"]
    override = _wants_update(user_text)

    # Email
    m = EMAIL_TEXT_RE.search(user_text)
    if m:
        email = (m.group(1) + "@" + m.group(2)).lower()
        _set_field(contact, "email", email, override=override or not contact.get("email"), label="email")

    # Phone
    allow_short = _looks_like_phone_utterance(user_text) or override
    for cand in INTL_PHONE_CANDIDATE.findall(user_text):
        norm = _normalize_numeric_candidate(cand, allow_short=allow_short)
        if norm:
            _set_field(contact, "phone", norm, override=override or not contact.get("phone"), label="phone")
            break
    if not contact.get("phone") or override:
        spoken = _spoken_to_digits(user_text, allow_short=allow_short)
        if spoken:
            _set_field(contact, "phone", spoken, override=override or not contact.get("phone"), label="phone")

    # Name
    for pat in NAME_PATTERNS:
        m = pat.search(user_text)
        if m:
            raw = m.group(1).strip()
            name = _clean_name(raw)
            if name:
                _set_field(contact, "name", name, override=override or not contact.get("name"), label="name")
                break

    # Save to Excel once per unique pair
    try:
        name = contact.get("name")
        phone = contact.get("phone")
        if name and phone:
            last = call_state.get("meta", {}).get("last_saved_pair")
            pair = (name, phone)
            if last != pair:
                _append_record_excel(name, phone)
                call_state["meta"]["last_saved_pair"] = pair
    except Exception as e:
        logger.error(f"[Excel] save attempt failed: {e}")

# ---- MINIMAL ADD: robust pretty address + recap helpers ----
def _pretty_address(addr: Dict[str, Optional[str]]) -> str:
    addr = addr or {}
    parts = []
    if addr.get("street") and addr.get("house"):
        parts.append(f"{addr['street']} {addr['house']}")
    elif addr.get("street"):
        parts.append(addr["street"])
    if addr.get("postal") and addr.get("city"):
        parts.append(f"{addr['postal']} {addr['city']}")
    elif addr.get("postal"):
        parts.append(addr["postal"])
    elif addr.get("city"):
        parts.append(addr["city"])
    return ", ".join(parts) if parts else "(unknown)"

_RECAP_CUE = re.compile(
    r"\b(recap|repeat|summary|summarize|show (?:all )?(?:my|the) (?:info|details)|"
    r"what (?:info|details) (?:do you|you) (?:have|know)|"
    r"appointment details|repeat the details|repeat info|"
    r"zusammenfassung|wiederhol|alle details|meine daten|was weißt du über mich|"
    r"termin details|termindaten)\b",
    re.IGNORECASE
)

def _wants_recap(text: str) -> bool:
    return bool(_RECAP_CUE.search(text or ""))

def format_known_info(call_state: dict) -> str:
    c = call_state.get("contact", {})
    b = call_state.get("booking", {})
    addr = _pretty_address(c.get("address") or {})
    lines = [
        "Appointment Summary / Termin-Zusammenfassung",
        "",
        "Contact / Kontakt:",
        f"• Name: {c.get('name') or '(unknown)'}",
        f"• Phone: {c.get('phone') or '(unknown)'}",
        f"• Email: {c.get('email') or '(unknown)'}",
        "",
        "Booking / Termin:",
        f"• Vehicle/Service: {b.get('vehicle') or '(unknown)'}",
        f"• Problem/Description: {b.get('problem') or '(unknown)'}",
        f"• Slot: {(b.get('slot_start') or '(unknown)')} – {(b.get('slot_end') or '(unknown)')}",
    ]
    return "\n".join(lines)
# ------------------------------------------------------------

# --- EMAIL ADD: confirmation + summary detectors ---
_CONFIRM_CUE = re.compile(
    r"\b(appointment\s+confirmed|confirmed\s+appointment|your\s+appointment\s+is\s+confirmed|"
    r"termin\s+bestätigt|termin\s+ist\s+bestätigt|ich\s+habe\s+den\s+termin\s+eingetragen|"
    r"termin\s+gebucht|termin\s+vereinbart)\b",
    re.IGNORECASE
)
_SUMMARY_CUE = re.compile(
    r"(?:^|\n)\s*zusammenfassung\s*:|(?:^|\n)\s*appointment\s+summary\s*:|"
    r"(?:^|\n)\s*termin[-\s]*zusammenfassung\s*:",
    re.IGNORECASE
)
_SUMMARY_BULLETS = re.compile(
    r"(?:^|\n)\s*-\s*(fahrzeug|vehicle)\s*:\s*.+"
    r".*?(?:^|\n)\s*-\s*(problem|beschreibung|issue|description)\s*:\s*.+"
    r".*?(?:^|\n)\s*-\s*(termin|slot|time)\s*:\s*.+"
    r".*?(?:^|\n)\s*-\s*(name)\s*:\s*.+"
    r".*?(?:^|\n)\s*-\s*(telefon|phone)\s*:\s*.+"
    r".*?(?:^|\n)\s*-\s*(e-?mail)\s*:\s*.+",
    re.IGNORECASE | re.DOTALL
)
# NEW: exact “Zusammenfassung” block extractor (verbatim)
_SUMMARY_BLOCK = re.compile(
    r'((?:^|\n)\s*(?:Zusammenfassung|Termin[-\s]*Zusammenfassung|Appointment\s+Summary)\s*:\s*\n(?:\s*(?:[-•]\s.*)\n?)+)',
    re.IGNORECASE
)
# ---------------------------------------------------

def _maybe_send_confirmation_email(assistant_text: str, call_state: dict):
    try:
        txt = assistant_text or ""
        # Trigger when explicit confirmation OR a recognizable summary header/bullets appear
        if not (_CONFIRM_CUE.search(txt) or _SUMMARY_CUE.search(txt) or _SUMMARY_BULLETS.search(txt)):
            return

        meta = call_state.setdefault("meta", {})
        if meta.get("email_sent"):
            return  # remove if you want multiple sends per session

        email_to = (call_state.get("contact") or {}).get("email")
        if not email_to:
            logger.info("[EMAIL] No recipient email captured; skipping send.")
            return

        # Extract the exact Zusammenfassung block (verbatim). If missing, skip sending.
        m = _SUMMARY_BLOCK.search(txt)
        if not m:
            logger.info("[EMAIL] No explicit 'Zusammenfassung' block found in assistant text; skipping email.")
            return

        subject = f"{state._company_name()} — Terminbestätigung / Appointment Confirmation"
        body = m.group(1).strip()  # send the exact block as-is

        _send_email_appointment_summary(email_to, subject, body)
        meta["email_sent"] = True
        logger.info(f"[EMAIL] Confirmation sent to {email_to}")
    except Exception as e:
        logger.error(f"[EMAIL] Failed to maybe-send confirmation: {e}")

def _send_email_appointment_summary(to_email: str, subject: str, body: str):
    server = os.getenv("SMTP_SERVER", "").strip()
    port_str = os.getenv("SMTP_PORT", "587").strip()
    email_from = os.getenv("EMAIL_FROM", "").strip()
    password = os.getenv("EMAIL_PASSWORD", "").strip()

    if not (server and port_str and email_from and password and to_email):
        logger.error("[EMAIL] Missing SMTP env vars or recipient.")
        return

    try:
        port = int(port_str)
    except Exception:
        port = 587

    msg = EmailMessage()
    msg["From"] = email_from
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    try:
        if port == 465:
            with smtplib.SMTP_SSL(server, port, timeout=15) as s:
                s.login(email_from, password)
                s.send_message(msg)
        else:
            with smtplib.SMTP(server, port, timeout=15) as s:
                s.ehlo()
                try:
                    s.starttls()
                except Exception:
                    pass
                s.login(email_from, password)
                s.send_message(msg)
    except Exception as e:
        logger.error(f"[EMAIL] SMTP send failed: {e}")

def build_system_prompt_with_memory(call_state: dict) -> str:
    """
    Build the system prompt that gets sent to the LLM every turn.
    This includes:
    - company/business info injected from env
    - known caller contact info
    - known booking info
    - behavior rules (don't ask again, etc.)
    """
    base = state.base_system_prompt or ""

    name = call_state["contact"].get("name")
    phone = call_state["contact"].get("phone")
    greeted = call_state["meta"].get("greeted")

    addr = call_state["contact"].get("address", {})
    parts = []
    if addr.get("street") and addr.get("house"):
        parts.append(f"{addr['street']} {addr['house']}")
    elif addr.get("street"):
        parts.append(addr["street"])
    if addr.get("postal") and addr.get("city"):
        parts.append(f"{addr['postal']} {addr['city']}")
    elif addr.get("postal"):
        parts.append(addr["postal"])
    elif addr.get("city"):
        parts.append(addr["city"])
    pretty_addr = ", ".join(parts) if parts else "(unknown)"

    # NEW: booking section
    booking = call_state.get("booking", {})
    vehicle = booking.get("vehicle") or "(unknown)"
    problem = booking.get("problem") or "(unknown)"
    slot_start = booking.get("slot_start") or "(unknown)"
    slot_end   = booking.get("slot_end")   or "(unknown)"

    # explicit recap rule
    recap_rule = (
        "- Wenn der Nutzer um eine **Wiederholung/Zusammenfassung** bittet "
        "(z. B. „repeat“, „recap“, „summary“, „Termin details“), "
        "dann liste **alle bekannten** Kontaktdaten und Terminangaben als kurze Aufzählung auf. "
        "Das Wiederholen der eigenen Daten des Nutzers ist **erlaubt**.\n"
    )

    memory_lines = [
        "Call/Chat meta:",
        f"- Already greeted: {'yes' if greeted else 'no'}",
        "",
        "Known caller details:",
        f"- Name: {name}" if name else "- Name: (unknown)",
        f"- Phone: {phone}" if phone else "- Phone: (unknown)",
        f"- Address: {pretty_addr}",
        "",
        "Known booking details:",
        f"- Vehicle/Service: {vehicle}",
        f"- Problem: {problem}",
        f"- Slot: {slot_start} bis {slot_end}",
        "",
        "Instructions for you (the assistant):",
        "- If a detail above is known, DO NOT ask for it again. Just confirm it naturally.",
        "- If a detail is unknown and becomes relevant, ask politely ONCE.",
        "- Do not re-greet if already greeted.",
        "- If booking details exist (vehicle/problem/slot), prefer confirming and summarizing instead of re-asking.",
        "- Toward completion, offer a brief, natural closing check-in.",
        recap_rule,
    ]

    return (base + "\n" if base else "") + "\n".join([m for m in memory_lines if m is not None]) + "\n"

# ---------- LLM streaming for voice ----------
_SENTENCE_END = re.compile(r'([.!?])(\s|$)')
_CLAUSE_END   = re.compile(r'([,;])(\s|$)')

def _split_complete_sentences(buf: str):
    out = []
    i = 0
    for m in _SENTENCE_END.finditer(buf):
        j = m.end()
        s = buf[i:j].strip()
        if s:
            out.append(s)
        i = j
    return out, buf[i:]

async def llm_stream_sentences(history, user_text, call_state):
    if not OPENAI_API_KEY:
        yield "Entschuldigung, mein KI-Gehirn ist offline."
        return
    system_prompt = build_system_prompt_with_memory(call_state)
    messages = [{"role": "system", "content": system_prompt}] + history[-12:] + [{"role": "user", "content": user_text}]
    try:
        async with httpx_client.stream(
            "POST", "https://api.openai.com/v1/chat/completions",
            json={
                "model": "gpt-4o-mini",
                "messages": messages,
                "temperature": 0.3,
                "max_tokens": 320,
                "stream": True
            },
        ) as r:
            r.raise_for_status()
            buf = ""
            last_flush = time.perf_counter()
            MAX_WAIT = 1.2
            MIN_CHARS = 40
            async for line in r.aiter_lines():
                if not line or not line.startswith("data:"):
                    continue
                data = line[5:].strip()
                if data == "[DONE]":
                    break
                try:
                    obj = json.loads(data)
                except Exception:
                    continue
                delta = obj.get("choices", [{}])[0].get("delta", {})
                token = delta.get("content")
                if token:
                    buf += token
                    sentences, buf = _split_complete_sentences(buf)
                    for s in sentences:
                        yield s
                        last_flush = time.perf_counter()
                    i = 0
                    for m in _CLAUSE_END.finditer(buf):
                        j = m.end()
                        s = buf[i:j].strip()
                        if s:
                            yield s
                            last_flush = time.perf_counter()
                        i = j
                    if i:
                        buf = buf[i:]
                    now = time.perf_counter()
                    if len(buf) >= MIN_CHARS and (now - last_flush) > MAX_WAIT:
                        yield buf.strip()
                        buf = ""
                        last_flush = now
            rem = buf.strip()
            if rem:
                yield rem
    except Exception as e:
        logger.error(f"LLM stream error: {e}")
        yield "Verstanden."

# ---------- Goodbye detection ----------
_GOODBYE_PAT = re.compile(
    r"\b(good\s*bye|goodbye|bye|ciao)\b|have a nice day\b|have a wonderful day\b|have a great day\b|"
    r"Tschüss\b|bis später\b|bis bald\b|auf wiederhören\b|auf wiedersehen\b|Verabschiedung\b|schönen Tag",
    re.I
)
def says_goodbye(text: str) -> bool:
    return bool(_GOODBYE_PAT.search(text or ""))

# ====== TIME/DATE/ETC (SSML helpers) ======
DATE_RE   = re.compile(r'\b(20\d{2}|19\d{2})[-/.](0?[1-9]|1[0-2])[-/.](0?[1-9]|[12]\d|3[01])\b')
TEL_RE    = re.compile(r'(?<!\w)(\+?\d[\d\s().\-]{4,}\d)(?!\w)')
CUR_RE    = re.compile(r'(\b\d{1,3}(?:[.,]\d{3})*(?:[.,]\d+)?\s?(?:€|EUR|€\s?|EUR\b))|((?:€|EUR)\s?\d+(?:[.,]\d+)?)', re.IGNORECASE)
URL_RE    = re.compile(r'\bhttps?://\S+|\bwww\.\S+', re.I)
EMAIL_SPLIT_RE = re.compile(r'\b([A-Za-z0-9._%+\-]+)@([A-Za-z0-9.\-]+\.[A-Za-z]{2,})\b')
MONTHS_DE = ("Januar","Februar","März","April","Mai","Juni","Juli","August","September","Oktober","November","Dezember")
MONTHS_RX = r'(?:' + '|'.join(MONTHS_DE) + r')'
ORDINAL_DATE_RE = re.compile(r'\b(den|am)?\s*(\d{1,2})\.\s+(' + MONTHS_RX + r')\b', re.IGNORECASE)
ZB_RE = re.compile(r'\bz\s*\.?\s*B\s*\.?\b', re.IGNORECASE)

def _wrap_email_local_spell(m):
    local, domain = m.group(1), m.group(2)
    return f'<say-as interpret-as="characters">{html.escape(local)}</say-as> ät {html.escape(domain)}'
def _fix_zb(_m):
    return '<sub alias="zum Beispiel">z. B.</sub>'

_WS = r'[\s\u00A0\u2009\u202F]*'
_TIME = rf'(?<!\d)([01]?\d|2[0-3]){_WS}[:.]{_WS}([0-5]\d)(?:{_WS}[:.]{_WS}([0-5]\d))?(?!\d)'
TIME_RE = re.compile(_TIME)
TIME_RANGE_RE = re.compile(
    rf'(?<!\d)(?:\bvon{_WS})?(?P<h1>[01]?\d|2[0-3]){_WS}[:.]{_WS}(?P<m1>[0-5]\d)(?:{_WS}[:.]{_WS}(?P<s1>[0-5]\d))?'
    rf'{_WS}(?:–|—|-|\bbis\b|\bbus\b){_WS}'
    rf'(?P<h2>[01]?\d|2[0-3]){_WS}[:.]{_WS}(?P<m2>[0-5]\d)(?:{_WS}[:.]{_WS}(?P<s2>[0-5]\d))?'
    rf'(?:{_WS}Uhr\b)?(?!\d)', re.IGNORECASE
)
TIME_WITH_UHR_RE = re.compile(
    rf'(?<!\d)(?P<h>[01]?\d|2[0-3]){_WS}[:.]{_WS}(?P<m>[0-5]\d)(?:{_WS}[:.]{_WS}(?P<s>[0-5]\d))?{_WS}Uhr\b(?!\d)',
    re.IGNORECASE
)

_HOUR = [
    "null","ein","zwei","drei","vier","fünf","sechs","sieben","acht","neun","zehn","elf","zwölf",
    "dreizehn","vierzehn","fünfzehn","sechzehn","siebzehn","achtzehn","neunzehn",
    "zwanzig","einundzwanzig","zweiundzwanzig","dreiundzwanzig","vierundzwanzig"
]
_ONES = ["null","eins","zwei","drei","vier","fünf","sechs","sieben","acht","neun"]
_TENS = ["","zehn","zwanzig","dreißig","vierzig","fünfzig"]
_SPECIAL = {
    10:"zehn",11:"elf",12:"zwölf",13:"dreizehn",14:"vierzehn",15:"fünfzehn",
    16:"sechzehn",17:"achtzehn",19:"neunzehn"
}

def _min_words(n: int) -> str:
    if n == 0:
        return ""
    if n < 10:
        return _ONES[n]
    if 10 <= n < 20:
        return _SPECIAL[n]
    if n % 10 == 0:
        return _TENS[n // 10]
    return f"{_ONES[n % 10]}und{_TENS[n // 10]}"

def _time_words(h: int, m: int, s: int) -> str:
    hour = "ein" if h == 1 else _HOUR[h]
    base = f"{hour} Uhr"
    if m == 0 and s == 0:
        return base
    mins = _min_words(m)
    if s:
        secs = _min_words(s)
        return f"{base} {mins} {secs} Sekunden" if mins else f"{base} {secs} Sekunden"
    return f"{base} {mins}" if mins else base

def _canon_int(x: Optional[str]) -> int:
    try:
        return int(x or "0")
    except:
        return 0

def _sub_time_single(m: re.Match) -> str:
    H = _canon_int(m.group(1)); M = _canon_int(m.group(2)); S = _canon_int(m.group(3))
    return _time_words(H, M, S)

def _sub_time_with_uhr(m: re.Match) -> str:
    H = _canon_int(m.group('h')); M = _canon_int(m.group('m')); S = _canon_int(m.group('s'))
    return _time_words(H, M, S)

def _sub_time_range(m: re.Match) -> str:
    H1 = _canon_int(m.group('h1')); M1 = _canon_int(m.group('m1')); S1 = _canon_int(m.group('s1'))
    H2 = _canon_int(m.group('h2')); M2 = _canon_int(m.group('m2')); S2 = _canon_int(m.group('s2'))
    return f"{_time_words(H1, M1, S1)} bis {_time_words(H2, M2, S2)}"

def _strip_markdown(text: str) -> str:
    text = re.sub(r'\*{1,3}([^\*\n]+)\*{1,3}', r'\1', text)
    text = re.sub(r'_{1,3}([^_\n]+)_{1,3}', r'\1', text)
    text = re.sub(r'`{1,3}([^`\n]+)`{1,3}', r'\1', text)
    return text

def _normalize_bullets(text: str) -> str:
    def repl(line): return re.sub(r'^\s*(?:[-*•]\s+|\d+\.\s+)', '• ', line)
    return "\n".join(repl(l) for l in text.splitlines())

def auto_ssml(text: str, lang="de-DE", voice=None) -> str:
    voice = voice or os.getenv("AZURE_TTS_VOICE", "de-DE-AmalaNeural")
    rate = os.getenv("AZURE_TTS_RATE", "+25%")
    gap_ms = int(os.getenv("AZURE_TTS_SENTENCE_GAP_MS", "40"))
    if not text:
        text = ""
    t = unicodedata.normalize("NFKC", text).replace("\u00A0"," ").replace("\u2009"," ").replace("\u202F"," ")
    t = _strip_markdown(t)
    t = _normalize_bullets(t)
    t = ZB_RE.sub(_fix_zb, t)
    t = TIME_RANGE_RE.sub(_sub_time_range, t)
    t = TIME_WITH_UHR_RE.sub(_sub_time_with_uhr, t)
    t = TIME_RE.sub(_sub_time_single, t)
    t = DATE_RE.sub(lambda m: f'<say-as interpret-as="date" format="dd-mm-yyyy">{m.group(0)}</say-as>', t)
    t = TEL_RE.sub(lambda m: f'<say-as interpret-as="telephone">{html.escape(m.group(0))}</say-as>', t)
    t = CUR_RE.sub(lambda m: f'<say-as interpret-as="currency">{html.escape(m.group(0).replace(" ", ""))}</say-as>', t)
    t = URL_RE.sub(lambda m: f'<say-as interpret-as="characters">{html.escape(m.group(0))}</say-as>', t)
    t = EMAIL_SPLIT_RE.sub(_wrap_email_local_spell, t)
    t = ORDINAL_DATE_RE.sub(
        lambda m: f'{(m.group(1) or "").strip()+" " if (m.group(1) or "").strip() else ""}'
                  f'<sub alias="{_de_ordinal_day(int(m.group(2)))}">{m.group(2)}.</sub> {m.group(3)}',
        t
    )
    t = re.sub(r'(?<!\d):\s', ' — ', t)
    sentences = [s for s in re.split(r'(?<=[.!?])\s+', t) if s]
    body = f"<s>{t}</s>" if not sentences else '<break time="{0}ms"/>'.format(gap_ms).join(
        f"<s>{s}</s>" for s in sentences
    )
    silence = f'''
      <mstts:silence type="Leading" value="0ms"/>
      <mstts:silence type="SentenceBoundary" value="{max(0, min(gap_ms, 60))}ms"/>
      <mstts:silence type="Tailing" value="0ms"/>
    '''
    return f'''<speak version="1.0" xml:lang="{lang}" xmlns:mstts="https://www.w3.org/2001/mstts">
  <voice name="{voice}">
    <prosody rate="{rate}">
      {silence}
      {body}
    </prosody>
  </voice>
</speak>'''

def _de_ordinal_day(n: int) -> str:
    irregular = {1: "ersten", 3: "dritten", 7: "siebten", 8: "achten"}
    return irregular.get(n, f"{n}sten" if (n in (0,6,9) or (n >= 20 and n % 10 == 0)) else f"{n}ten")

# ---------- Helper: default WS URL on Azure ----------
def _default_ws_url() -> Optional[str]:
    try:
        host = request.headers.get("Host") or request.host
        if not host:
            return None
        return f"wss://{host}/media"
    except Exception:
        return None

# ---------- TwiML entry (VOICE CALLS) ----------
@app.post("/incoming-call")
async def incoming_call():
    ws_url = os.getenv("MEDIA_WS_URL") or _default_ws_url()
    if not ws_url:
        return "Missing MEDIA_WS_URL", 500
    form = await request.form
    call_sid = form.get("CallSid")
    logger.info(f"Call start: {call_sid} -> streaming to {ws_url}")
    vr = VoiceResponse()
    connect = vr.connect()
    connect.stream(url=ws_url)
    return str(vr)

# ---------- Tiny TwiML for forced hangup (fallback) ----------
@app.get("/hangup-twiml")
async def hangup_twiml():
    vr = VoiceResponse()
    vr.hangup()
    return str(vr)

# ---------- Robust REST hangup ----------
async def _hangup_via_twilio_rest(call_sid: str, app_base_url: str) -> bool:
    acc = (os.getenv("TWILIO_ACCOUNT_SID") or "").strip()
    tok = (os.getenv("TWILIO_AUTH_TOKEN") or "").strip()
    api_key = (os.getenv("TWILIO_API_KEY_SID") or "").strip()
    api_secret = (os.getenv("TWILIO_API_KEY_SECRET") or "").strip()

    if not acc or not call_sid:
        logger.error(f"[HANGUP] Missing TWILIO_ACCOUNT_SID or callSid (acc set? {bool(acc)}, callSid set? {bool(call_sid)})")
        return False

    if api_key and api_secret:
        auth = (api_key, api_secret); auth_mode = "api_key"
    elif tok:
        auth = (acc, tok); auth_mode = "auth_token"
    else:
        logger.error("[HANGUP] No Twilio credentials found. Provide TWILIO_AUTH_TOKEN or API KEY/SECRET.")
        return False

    base = f"https://api.twilio.com/2010-04-01/Accounts/{acc}/Calls/{call_sid}.json"
    async with httpx.AsyncClient(timeout=8.0) as c:
        # Try direct hangup
        try:
            r1 = await c.post(base, data={"Status": "completed"}, auth=auth)
            if r1.status_code // 100 == 2:
                logger.info(f"[HANGUP] Completed via Status=completed using {auth_mode}.")
                return True
            logger.warning(f"[HANGUP] Direct hangup failed ({r1.status_code}): {r1.text[:300]}")
        except Exception as e:
            logger.error(f"[HANGUP] Direct hangup error: {e}")

        # Fallback: redirect to /hangup-twiml
        try:
            hangup_url = f"{app_base_url}/hangup-twiml"
            r2 = await c.post(base, data={"Url": hangup_url, "Method": "GET"}, auth=auth)
            if r2.status_code // 100 == 2:
                logger.info(f"[HANGUP] Redirected to {hangup_url} (Twiml <Hangup/>) using {auth_mode}.")
                return True
            if r2.status_code == 401:
                logger.error(
                    "[HANGUP] 401 Unauthorized from Twilio.\n"
                    " • TWILIO_ACCOUNT_SID must match the (sub)account that owns the CallSid.\n"
                    " • API keys or Auth Token must be for that SAME (sub)account.\n"
                    " • Ensure live creds, no trailing spaces.\n"
                )
            else:
                logger.error(f"[HANGUP] Redirect failed ({r2.status_code}): {r2.text[:300]}")
        except Exception as e:
            logger.error(f"[HANGUP] Redirect hangup error: {e}")

    return False

# ---------- WebSocket media loop (VOICE REALTIME) ----------
@app.websocket("/media")
async def media():
    subs = getattr(websocket, "subprotocols", []) or []
    if "audio" in subs:
        await websocket.accept(subprotocol="audio")
    else:
        await websocket.accept()
    logger.info("WebSocket connected.")
    stream_sid = None

    # Build base URL for TwiML redirect fallback
    app_base_url = os.getenv("APP_BASE_URL")
    if not app_base_url:
        host = (websocket.headers.get("Host") or os.getenv("WEBSITE_HOSTNAME") or "").strip()
        if not host:
            logger.warning("[WS] Could not determine host for app_base_url; set APP_BASE_URL in env.")
            app_base_url = "https://localhost"
        else:
            app_base_url = f"https://{host}"
    logger.info(f"[WS] app_base_url set to {app_base_url}")

    # Per-call state
    history = []
    greeted = False

    call_state = {
        "contact": {
            "name": None,
            "phone": None,
            "email": None,
            "address": {"street": None, "house": None, "postal": None, "city": None}
        },
        "booking": {  # NEW booking memory so voice and WhatsApp share same schema
            "vehicle": None,
            "problem": None,
            "slot_start": None,
            "slot_end": None
        },
        "meta": {"greeted": False, "last_saved_pair": None}
    }

    # Turn control
    current_turn = {"id": 0}
    llm_task: Optional[asyncio.Task] = None

    # TTS pipeline
    tts_queue: asyncio.Queue = asyncio.Queue()
    tts_busy = False
    tts_cancel = False
    last_tts_start_ms = 0.0
    FRAME = 160  # 20 ms @ 8k μ-law

    # Barge-in
    speech_streak_frames = 0
    REQ_STREAK_FRAMES = 5

    # Other state
    POST_TTS_GUARD_MS = int(os.getenv("POST_TTS_GUARD_MS", "1500"))
    last_user_media_ms = time.time() * 1000.0
    last_tts_end_ms = time.time() * 1000.0
    last_bot_asked_question = False
    last_assistant_sentence_ms = time.time() * 1000.0
    interaction_started = False
    allow_hangup = False  # hangup allowed only after caller talks

    smooth_task: Optional[asyncio.Task] = None
    call_sid_for_rest: Optional[str] = None
    end_called = False

    # Azure ASR
    recognizer, push_stream = state.make_asr()
    loop = asyncio.get_event_loop()

    final_queue: asyncio.Queue = asyncio.Queue()
    def on_recognized(evt: speechsdk.SpeechRecognitionEventArgs):
        if evt.result.reason == speechsdk.ResultReason.RecognizedSpeech and evt.result.text:
            txt = evt.result.text.strip()
            if txt:
                loop.call_soon_threadsafe(final_queue.put_nowait, txt)
    recognizer.recognized.connect(on_recognized)
    recognizer.start_continuous_recognition_async()

    # Create ONE synthesizer per call
    synthesizer = state.make_tts()
    current_chunk_q: Optional[asyncio.Queue] = None

    def _tts_on_synth(evt: speechsdk.SpeechSynthesisEventArgs):
        try:
            if evt and evt.result and evt.result.audio_data and current_chunk_q is not None:
                loop.call_soon_threadsafe(current_chunk_q.put_nowait, evt.result.audio_data)
        except Exception as e:
            logger.error(f"TTS on_synth error: {e}")

    def _tts_on_completed(_evt):
        try:
            if current_chunk_q is not None:
                loop.call_soon_threadsafe(current_chunk_q.put_nowait, None)
        except Exception:
            pass

    synthesizer.synthesizing.connect(_tts_on_synth)
    synthesizer.synthesis_completed.connect(_tts_on_completed)
    synthesizer.synthesis_canceled.connect(_tts_on_completed)

    async def tts_worker():
        nonlocal tts_busy, tts_cancel, last_tts_start_ms, last_tts_end_ms, stream_sid, current_chunk_q, allow_hangup
        while True:
            item = await tts_queue.get()
            if item is None:
                break
            item_turn, text = item
            if item_turn != current_turn["id"]:
                continue
            try:
                tts_busy = True
                tts_cancel = False
                last_tts_start_ms = time.time() * 1000.0
                if getattr(websocket, "closed", False):
                    break
                chunk_q: asyncio.Queue = asyncio.Queue()
                current_chunk_q = chunk_q
                ssml = auto_ssml(text, lang="de-DE", voice=os.getenv("AZURE_TTS_VOICE"))
                synth_coro = asyncio.to_thread(synthesizer.speak_ssml_async(ssml).get)
                next_tick = time.perf_counter()
                sent_frames = 0
                while True:
                    if tts_cancel or getattr(websocket, "closed", False):
                        try:
                            await websocket.send(json.dumps({"event": "clear", "streamSid": stream_sid}))
                        except Exception:
                            pass
                        break
                    chunk = await chunk_q.get()
                    if chunk is None:
                        break
                    i, n = 0, len(chunk)
                    while not tts_cancel and i < n and not getattr(websocket, "closed", False):
                        frame = chunk[i:i+FRAME]
                        i += len(frame)
                        if not frame:
                            break
                        payload = base64.b64encode(frame).decode("ascii")
                        try:
                            await websocket.send(json.dumps({
                                "event": "media",
                                "streamSid": stream_sid,
                                "media": {"payload": payload}
                            }))
                        except Exception:
                            tts_cancel = True
                            break
                        sent_frames += 1
                        next_tick += 0.02
                        delay = next_tick - time.perf_counter()
                        if delay > 0:
                            await asyncio.sleep(delay)
                logger.info(f"[TTS-streaming] frames sent: {sent_frames}")
                try:
                    await synth_coro
                except Exception:
                    pass
                finally:
                    current_chunk_q = None
                last_tts_end_ms = time.time() * 1000.0
                if not getattr(websocket, "closed", False) and item_turn == current_turn["id"]:
                    try:
                        await websocket.send(json.dumps({
                            "event": "mark",
                            "streamSid": stream_sid,
                            "mark": {"name": "tts-end"}
                        }))
                    except Exception:
                        pass

                if says_goodbye(text) and item_turn == current_turn["id"]:
                    if allow_hangup:
                        logger.info("[HANGUP] Goodbye phrase spoken; ending call (armed).")
                        await _end_call_and_close()
                        break
                    else:
                        logger.info("[HANGUP] Goodbye phrase in bot output ignored (not armed yet).")
            except Exception as e:
                logger.error(f"TTS worker error: {e}")
            finally:
                tts_busy = False
    asyncio.create_task(tts_worker())

    async def speak_llm_stream(turn_id: int, user_text: str):
        nonlocal last_bot_asked_question, interaction_started, last_assistant_sentence_ms
        assistant_accum = []
        async for sentence in llm_stream_sentences(history, user_text, call_state):
            if turn_id != current_turn["id"]:
                break
            s = sentence.strip()
            if not s:
                continue
            assistant_accum.append(s)
            interaction_started = True
            last_bot_asked_question = s.endswith("?")
            last_assistant_sentence_ms = time.time() * 1000.0
            await tts_queue.put((turn_id, s))
        full = " ".join(assistant_accum).strip()
        if assistant_accum and not assistant_accum[-1].endswith("?"):
            last_bot_asked_question = False
        if full and turn_id == current_turn["id"]:
            history.append({"role": "assistant", "content": full})
            logger.info(f"[Bot] {full}")
            # --- EMAIL ADD: trigger after confirmation/summary is spoken ---
            _maybe_send_confirmation_email(full, call_state)
            # --------------------------------------------------------------

    async def consume_finals():
        nonlocal llm_task, tts_cancel, interaction_started, allow_hangup
        while True:
            user_text = await final_queue.get()
            allow_hangup = True

            # Update memory every user final
            extract_contact(user_text, call_state)
            extract_address(user_text, call_state)

            current_turn["id"] += 1
            turn_id = current_turn["id"]
            interaction_started = True
            logger.info(f"[Caller] {user_text} | MEM: {call_state['contact']} | hangup_armed={allow_hangup}")

            if llm_task and not llm_task.done():
                llm_task.cancel()
                try:
                    await llm_task
                except:
                    pass

            tts_cancel = True
            try:
                while True:
                    _ = tts_queue.get_nowait()
            except asyncio.QueueEmpty:
                pass

            history.append({"role": "user", "content": user_text})
            llm_task = asyncio.create_task(speak_llm_stream(turn_id, user_text))
    asyncio.create_task(consume_finals())

    async def _hangup_via_rest_or_redirect():
        if call_sid_for_rest:
            ok = await _hangup_via_twilio_rest(call_sid_for_rest, app_base_url)
            if not ok:
                logger.error("[HANGUP] REST/Redirect hangup did not succeed; closing WS anyway (Twilio may keep call).")

    async def _end_call_and_close():
        nonlocal end_called
        if end_called:
            return
        end_called = True
        await _hangup_via_rest_or_redirect()
        try:
            await websocket.close()
        except Exception:
            pass

    # WS main loop
    try:
        while True:
            raw = await websocket.receive()
            if raw is None:
                break
            msg = json.loads(raw)
            event = msg.get("event")

            if event == "start":
                stream_sid = msg["start"]["streamSid"]
                call_sid_for_rest = ((msg.get("start") or {}).get("callSid") or msg.get("callSid"))
                logger.info(f"Stream started: {stream_sid} (callSid={call_sid_for_rest})")
                if not greeted:
                    greeted = True
                    call_state["meta"]["greeted"] = True
                    current_turn["id"] += 1
                    greeting_text = _env_text("GREETING")
                    if greeting_text and greeting_text.strip():
                        await tts_queue.put((current_turn["id"], greeting_text))
                    else:
                        logger.warning("GREETING not set in environment; skipping initial spoken greeting.")

            elif event == "media":
                b64 = msg["media"]["payload"]
                pcm16 = mulaw_b64_to_pcm16_bytes(b64)
                await asyncio.to_thread(push_stream.write, pcm16)

                if is_speech(pcm16):
                    speech_streak_frames = min(10, speech_streak_frames + 1)
                else:
                    speech_streak_frames = 0

                if tts_busy and speech_streak_frames >= 5:
                    tts_cancel = True
                    try:
                        await websocket.send(json.dumps({"event": "clear", "streamSid": stream_sid}))
                    except Exception:
                        pass
                    try:
                        while True:
                            _ = tts_queue.get_nowait()
                    except asyncio.QueueEmpty:
                        pass

            elif event == "stop":
                logger.info("Stream stopped by Twilio.")
                break

    except Exception as e:
        logger.error(f"WS error: {e}")
    finally:
        try:
            if llm_task and not llm_task.done():
                llm_task.cancel()
                try:
                    await llm_task
                except:
                    pass
            await tts_queue.put(None)
            if smooth_task and not smooth_task.done():
                smooth_task.cancel()
                try:
                    await smooth_task
                except:
                    pass
            push_stream.close()
            recognizer.stop_continuous_recognition_async()
        except Exception:
            pass
        logger.info("WebSocket closed.")

# ============================
# WHATSAPP MODE (text chat bot)
# ============================

# ---- NEW: persistent session storage on disk ----
SESSIONS_PATH = os.getenv("SESSIONS_PATH", "sessions_store.json")

def _load_sessions() -> Dict[str, Any]:
    try:
        if os.path.exists(SESSIONS_PATH):
            with open(SESSIONS_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    logger.info(f"[WA STORE] loaded {len(data)} sessions from disk")
                    return data
    except Exception as e:
        logger.error(f"[WA STORE] load failed: {e}")
    return {}

def _persist_sessions():
    try:
        tmp = SESSIONS_PATH + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(SESSIONS, f, ensure_ascii=False)
        os.replace(tmp, SESSIONS_PATH)
        logger.info("[WA STORE] sessions persisted to disk")
    except Exception as e:
        logger.error(f"[WA STORE] persist failed: {e}")

# global sessions dict (phone -> {history, call_state})
SESSIONS: Dict[str, Any] = _load_sessions()

async def llm_complete_once(history, user_text, call_state) -> str:
    """
    Non-streaming completion for WhatsApp.
    We reuse build_system_prompt_with_memory() so bot has
    the same memory rules.
    """
    if not OPENAI_API_KEY:
        return "Entschuldigung, mein KI-Gehirn ist offline."

    system_prompt = build_system_prompt_with_memory(call_state)

    messages = [{"role": "system", "content": system_prompt}] + history[-12:] + [
        {"role": "user", "content": user_text}
    ]

    try:
        resp = await httpx_client.post(
            "https://api.openai.com/v1/chat/completions",
            json={
                "model": "gpt-4o-mini",
                "messages": messages,
                "temperature": 0.3,
                "max_tokens": 320,
                "stream": False
            },
        )
        resp.raise_for_status()
        data = resp.json()
        txt = data["choices"][0]["message"]["content"].strip()
        return txt
    except Exception as e:
        logger.error(f"[WhatsApp LLM] error: {e}")
        return "Verstanden."

@app.post("/incoming-whatsapp")
async def incoming_whatsapp():
    """
    Twilio WhatsApp webhook.
    - identify sender
    - load/create their session
    - update memory (contact/address)
    - if user asks for a recap, return a deterministic summary
    - else LLM answer
    - respond with Twilio MessagingResponse
    - persist session to disk
    """
    form = await request.form
    from_number = form.get("From", "").strip()  # e.g. 'whatsapp:+49123456789'
    body_text   = form.get("Body", "").strip() or ""

    if not from_number:
        from_number = "unknown"

    # load or init session for this number
    sess = SESSIONS.get(from_number)
    if not sess:
        sess = {
            "history": [],
            "call_state": {
                "contact": {
                    "name": None,
                    "phone": None,
                    "email": None,
                    "address": {"street": None, "house": None, "postal": None, "city": None}
                },
                "booking": {  # same shape as voice
                    "vehicle": None,
                    "problem": None,
                    "slot_start": None,
                    "slot_end": None
                },
                "meta": {"greeted": False, "last_saved_pair": None}
            }
        }
        SESSIONS[from_number] = sess

        # first time: mark greeted so bot doesn't "Hallo willkommen" every message
        sess["call_state"]["meta"]["greeted"] = True

        # pre-store WhatsApp number as phone if empty
        phone_guess = re.sub(r"^whatsapp:", "", from_number)
        if phone_guess:
            _set_field(
                sess["call_state"]["contact"],
                "phone",
                phone_guess,
                override=False,
                label="phone"
            )

    # update memory with latest user text
    extract_contact(body_text, sess["call_state"])
    extract_address(body_text, sess["call_state"])

    # add user turn to history
    sess["history"].append({"role": "user", "content": body_text})
    logger.info(f"[WA User {from_number}] {body_text} | MEM: {sess['call_state']['contact']}")

    # deterministic recap when requested
    if _wants_recap(body_text):
        assistant_text = format_known_info(sess["call_state"])
    else:
        assistant_text = await llm_complete_once(sess["history"], body_text, sess["call_state"])

    # append assistant turn
    sess["history"].append({"role": "assistant", "content": assistant_text})
    logger.info(f"[WA Bot -> {from_number}] {assistant_text}")

    # EMAIL trigger when confirmation or summary text appears
    _maybe_send_confirmation_email(assistant_text, sess["call_state"])

    # persist session to disk so we don't forget name/address/booking next message
    _persist_sessions()

    # reply Twilio-style
    twiml_resp = MessagingResponse()
    twiml_resp.message(assistant_text)
    return str(twiml_resp)

# ---------- Shutdown ----------
@app.after_serving
async def _shutdown():
    try:
        await httpx_client.aclose()
    except Exception:
        pass

# ---------- Health ----------
@app.get("/")
async def home():
    return "AI Receptionist — Natural v12.3 (DE voice, fast responses, robust hangup with guard, + WhatsApp chat mode w/ persistent memory)"
